from urllib.parse import quote

import openpyxl
from rest_framework.views import APIView
from rest_framework.response import Response
from app01.models import College,Branch
from django.shortcuts import HttpResponse

# 给出branch和college的关系
class branch_college(APIView):
    def get(self, request):
        queryset = College.objects.all()
        college = []
        for obj in queryset:
            tmp = {
                "id": obj.pk,
                "label": obj.title,
            }
            college.append(tmp)
        branch = {}
        for x in college:
            tmps = []
            queryset = Branch.objects.filter(college_id=x['id']).all()
            for obj in queryset:
                tmp = {
                    "id": obj.pk,
                    "label": obj.title,
                    "college": obj.college_id,
                    "college_title": obj.college.title
                }
                tmps.append(tmp)
            branch[f"{x['label']}"] = tmps
        ans = {
            "college" : college,
            "branch" : branch,
        }
        return Response(ans)

# 从支部数据库导出excel
class branch_excel(APIView):
    def get(self,request):
        # django里面的特定调用数据库的方法
        queryset = Branch.objects.all()
        # 创建对象
        wb = openpyxl.Workbook()
        # 起文件名字
        filename = "branch.xlsx"
        # 创建sheet
        ws1 = wb.active
        ws1.title = "支部信息"
        # 给excel的单元格填写内容
        ws1['A1'].value = 'id'
        ws1['B1'].value = '名称'
        ws1['C1'].value = '所属学院'
        ws1['D1'].value = '所属学院id'
        row, col = 2, 1
        for obj in queryset:
            ws1.cell(row, col).value = obj.id   # 这里的字段一定要和数据表里对应
            ws1.cell(row, col+1).value = obj.title
            ws1.cell(row, col+2).value = obj.college.title
            ws1.cell(row, col+3).value = obj.college.id
            row = row + 1
        # print(66666)
        # 返回内容
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                            charset='utf-8')
        response['Content-Disposition'] = 'attachment; filename='+quote(str(filename))
        response['Access-Control-Expose-Headers'] = 'Content-Disposition'
        wb.save(response)  # 必须有
        return response
