import openpyxl
from app01.models import User,Branch
from django.shortcuts import HttpResponse
from rest_framework.response import Response
from rest_framework.status import HTTP_202_ACCEPTED,HTTP_200_OK,HTTP_404_NOT_FOUND

v = ("已上传", "未上传")

def user(request, college=None, branch=None, person=None):
    Chinese = request.data.get("Chinese")
    options = request.data.get("option")
    ans = [False, None]
    if not options:
        ans[1] = HttpResponse(status=202, content="找不到option参数")
        return ans
    if not Chinese:
        ans[1] = HttpResponse(status=202, content="找不到Chinese参数")
        return ans
    if len(options) != len(Chinese):
        ans[1] = HttpResponse(status=202, content="Chinese与options长度不一致")
        return ans
    op = {}
    # op["account"] = False   # 学号
    # op["name"] = False   # 姓名
    op["status"] = False  # 所处发展阶段
    op["gender"] = False  # 性别
    op["phone"] = False  # 电话
    op["classx"] = False  # 班级
    op["identity_card"] = False  # 身份证
    op["nation"] = False  # 民族
    op["native"] = False  # 籍贯
    op["college"] = False  # 所在学院
    op["branch"] = False  # 所在支部
    op["birth_date"] = False  # 出生日期,后端根据身份证写
    op["First_application"] = False  # 首次提交申请书时间
    op["semester"] = False  # 首次申请入党学期
    op["talking_date"] = False  # 党组织谈话时间
    op["talker"] = False  # 党组织谈话人
    #
    op["activists_date"] = False  # 列为入党积极分子时间
    op["courses_for_activists"] = False  # 入党积极分子培训班情况
    op["activists_graduation"] = False  # 入党积极分子培训班结业时间
    op["activists_grade"] = False  # 入党积极培训班成绩
    op["activists_appraising"] = False  # 入党积极分子评优情况
    op["activists_remarks"] = False  # 入党积极培训班情况备注
    op["contacts1"] = False  # 培养联系人1
    op["contacts2"] = False  # 培养联系人2
    op["investigation_registration"] = False  # 积极分子考察登记表是否上传
    #
    op["developer_date"] = False  # 列为发展对象的时间
    op["courses_for_developer"] = False  # 发展对象培训班情况
    op["developer_graduation"] = False  # 发展对象培训班结业时间
    op["developer_grade"] = False  # 发展对象培训班成绩
    op["developer_appraising"] = False  # 发展对象评优情况
    op["developer_remarks"] = False  # 发展对象培训班情况备注
    op["introducer1"] = False  # 入党介绍人1
    op["introducer2"] = False  # 入党介绍人2
    op["mass_discussion"] = False  # 群众座谈人员名单是否上传
    #
    op["probation_date"] = False  # 列为预备党员的时间\预备党员支部大会时间
    op["probation_remarks"] = False  # 预备党员情况备注
    op["number"] = False  # 入党志愿书编号
    op["formal_date"] = False  # 转为正式党员日期\党员支部大会时间
    #
    op["job"] = False  # 学生工作职位
    op["email"] = False  # 邮箱
    op["address"] = False  # 住址
    op["headmaster"] = False  # 班主任
    # 判断范围
    if person:
        queryset = User.objects.filter(pk=person)
    elif branch:
        queryset = User.objects.filter(branch_id=branch).all()
    elif college:
        queryset = User.objects.filter(college_id=college).all()
    else:
        queryset = User.objects.all()
    if not queryset.first():
        ans[1] = HttpResponse(status=404, content="Not found")
        return ans
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    # 给excel的单元格填写内容
    ws1['A1'].value = "学号"
    ws1['B1'].value = "姓名"
    for i in range(len(options)):
        ws1.cell(1, i + 3).value = Chinese[i]
        op[options[i]] = True
    row = 2
    for obj in queryset:
        col = 1
        ws1.cell(row, col).value = obj.account
        col += 1
        ws1.cell(row, col).value = obj.name
        col += 1
        if op["status"]:
            ws1.cell(row, col).value = obj.get_status_display()
            col += 1
        if op["gender"]:
            ws1.cell(row, col).value = obj.gender
            col += 1
        if op["phone"]:
            ws1.cell(row, col).value = obj.phone
            col += 1
        if op["classx"]:
            ws1.cell(row, col).value = obj.classx
            col += 1
        if op["identity_card"]:
            ws1.cell(row, col).value = obj.identity_card
            col += 1
        if op["nation"]:
            ws1.cell(row, col).value = obj.nation
            col += 1
        if op["native"]:
            ws1.cell(row, col).value = obj.native
            col += 1
        if op["college"]:
            ws1.cell(row, col).value = obj.college.title
            col += 1
        if op["branch"]:
            ws1.cell(row, col).value = obj.branch.title
            col += 1
        if op["birth_date"]:
            ws1.cell(row, col).value = obj.identity_card[6:14]
            col += 1
        if op["First_application"]:
            ws1.cell(row, col).value = obj.First_application
            col += 1
        if op["semester"]:
            ws1.cell(row, col).value = obj.semester
            col += 1
        if op["talking_date"]:
            ws1.cell(row, col).value = obj.talking_date
            col += 1
        if op["talker"]:
            ws1.cell(row, col).value = obj.talker
            col += 1
        if op["activists_date"]:
            ws1.cell(row, col).value = obj.activists_date
            col += 1
        if op["courses_for_activists"]:
            ws1.cell(row, col).value = obj.courses_for_activists
            col += 1
        if op["activists_graduation"]:
            ws1.cell(row, col).value = obj.activists_graduation
            col += 1
        if op["activists_grade"]:
            ws1.cell(row, col).value = obj.activists_grade
            col += 1
        if op["activists_appraising"]:
            ws1.cell(row, col).value = obj.activists_appraising
            col += 1
        if op["activists_remarks"]:
            ws1.cell(row, col).value = obj.activists_remarks
            col += 1
        if op["contacts1"]:
            ws1.cell(row, col).value = obj.contacts1
            col += 1
        if op["contacts2"]:
            ws1.cell(row, col).value = obj.contacts2
            col += 1
        if op["investigation_registration"]:
            if obj.investigation_registration:
                ws1.cell(row, col).value = v[0]
            else:
                ws1.cell(row, col).value = v[1]
            col += 1
        if op["developer_date"]:
            ws1.cell(row, col).value = obj.developer_date
            col += 1
        if op["courses_for_developer"]:
            ws1.cell(row, col).value = obj.courses_for_developer
            col += 1
        if op["developer_graduation"]:
            ws1.cell(row, col).value = obj.developer_graduation
            col += 1
        if op["developer_grade"]:
            ws1.cell(row, col).value = obj.developer_grade
            col += 1
        if op["developer_appraising"]:
            ws1.cell(row, col).value = obj.developer_appraising
            col += 1
        if op["developer_remarks"]:
            ws1.cell(row, col).value = obj.developer_remarks
            col += 1
        if op["introducer1"]:
            ws1.cell(row, col).value = obj.introducer1
            col += 1
        if op["introducer2"]:
            ws1.cell(row, col).value = obj.introducer2
            col += 1
        if op["mass_discussion"]:
            if obj.mass_discussion:
                ws1.cell(row, col).value = v[0]
            else:
                ws1.cell(row, col).value = v[1]
            col += 1
        if op["probation_date"]:
            ws1.cell(row, col).value = obj.probation_date
            col += 1
        if op["probation_remarks"]:
            ws1.cell(row, col).value = obj.probation_remarks
            col += 1
        if op["number"]:
            ws1.cell(row, col).value = obj.number
            col += 1
        if op["formal_date"]:
            ws1.cell(row, col).value = obj.formal_date
            col += 1
        if op["job"]:
            ws1.cell(row, col).value = obj.job
            col += 1
        if op["email"]:
            ws1.cell(row, col).value = obj.email
            col += 1
        if op["address"]:
            ws1.cell(row, col).value = obj.address
            col += 1
        if op["headmaster"]:
            ws1.cell(row, col).value = obj.headmaster
            col += 1
        row = row + 1
    ans[0], ans[1] = True, wb
    return ans

def impor(request,funcz,branch=None):
    if not branch:
        return Response(status=HTTP_404_NOT_FOUND)
    if not Branch.objects.filter(pk=branch).first():
        return Response(status=HTTP_404_NOT_FOUND)
    datae, flag = set(), False
    if request.FILES:
        file_obj = request.FILES.get("file_obj")
        if file_obj:
            try:
                wb = openpyxl.load_workbook(file_obj)
                ws = wb["Sheet1"]
                funcz(ws=ws, branch=int(branch), datae=datae)
            except Exception:
                return Response(status=HTTP_202_ACCEPTED, data={"entail": "不是.xlsx文件， 或者xlsx格式排版错误"})
        else:
            return Response(status=HTTP_202_ACCEPTED, data={"entail": "没有参数名为file_obj的文件"})
    else:
        return Response(status=HTTP_202_ACCEPTED,data={"entail": "There is no file."})
    if len(datae) == 0:
        flag = True
    return Response(status=HTTP_200_OK, data={"flag": flag, "error": datae})